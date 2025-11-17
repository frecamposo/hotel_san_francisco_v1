from django.shortcuts import render
from .i18n import *

from .models import *
# importar el modelo de tabla User 
from django.contrib.auth.models import User,Group
# importar librerias que permitan la validacion del login
from django.contrib.auth import authenticate,logout,login as login_aut
# importar libreria decoradora que permite evitar el ingreso de usuarios a las paginas web
from django.contrib.auth.decorators import login_required, permission_required

from django.shortcuts import redirect
from datetime import datetime, timedelta
import uuid
import math

from openpyxl import Workbook,load_workbook
from django.http import HttpResponse
from io import BytesIO

import os
import time

import qrcode
from io import BytesIO
from django.core.files import File
from PIL import Image,ImageDraw
from django.core.mail import EmailMessage
from django.http import HttpResponse
from django.core.files.base import ContentFile
import pandas as pd
# Create your views here.
import httpx
import asyncio
from datetime import datetime


documento_inf=[]
monto_buscar_inf=[]
diferencia_inf=[]
fecha_inf=[]
tc_inf=[]

informacion_proceso=[]

data_dolar=[]
async def obtener_datos():
    async with httpx.AsyncClient() as client:
        response = await client.get("https://mindicador.cl/api/dolar")  # Método GET
        response.raise_for_status()  # Lanza error si hay problema (4xx o 5xx)
        datos = response.json()      # Convierte respuesta JSON en dict de Python
        for item in datos['serie']:
            fecha_original = item["fecha"]
            fecha_obj = datetime.fromisoformat(fecha_original.replace("Z", "+00:00"))
            fecha_formateada = fecha_obj.strftime("%d-%m-%Y")
            #print(f"Fecha: {fecha_formateada}, Valor: {item['valor']}")
            fila={"fecha":fecha_formateada,"valor":item['valor']}
            data_dolar.append(fila)
        print(data_dolar)
        


def inicio(request):
    request.session["datos"]=""
    x={}

    # x["valor"]=request.session["datos"]
    # x["habitacion"]=habi(0)
    comentarios=Comentario.objects.all()
    mensaje={'comentarios':comentarios}
    return render(request,"index.html",mensaje)

def procesamiento(request):
    asyncio.run(obtener_datos())
    # definicion de variables que capturan las sumas por tipo de tarjeta
    diccionario = {}
    suma_amex=0
    suma_dinners=0
    suma_mc=0
    suma_master_card=0
    cant_master_card=0
    suma_visa=0
    i=0
    print("entro")
    dic_amex=[]
    contexto={}


    
    if request.method == 'POST':
    
        archivo1 = request.FILES.get('archivo1') # ERP
        archivo2 = request.FILES.get('archivo2') # amex
        archivo3 = request.FILES.get('archivo3') # dinners
        archivo4 = request.FILES.get('archivo4') # visa dolar
        archivo5 = request.FILES.get('archivo5') # mastercard
        archivo6 = request.FILES.get('archivo6') # banco
        archivo7 = request.FILES.get('archivo7') # transbank 
    
        abonado = 0
        
        if archivo6:
            print("ES VALIDO ARCHIVO 6")
            ws=0
            abono_dolar=pd.read_excel(archivo7,usecols=[1,2,3,4,5,6,7,8,9,10], nrows=80)   # transbank         
            for fila in abono_dolar.index:
                cuenta = abono_dolar.at[fila, 'Unnamed: 1']   
                #print(f"<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<< Cuenta leida: {cuenta}")
                if cuenta == "Abono Calculado (=):":
                    monto = abono_dolar.at[fila, 'Unnamed: 2']
                    #print(f"Monto Abono Calculado: {monto} a buscar")
                    contexto["msg"]=f"encontro abono por {monto}"
                    banco=pd.read_excel(archivo6,usecols=[1,2,3,4,5,6,7,8], nrows=1500)   # transbank                 
                    find_banco = banco[banco['Unnamed: 8'] == monto]
                    #print(f"Filas encontradas en banco para el monto {monto}: {len(find_banco)}")
                    if len(find_banco)>0:
                        #print(f"Encontrado el abono en el BANCO por el monto: {monto} **************")                            
                        abonado = monto
                        contexto['abonado']=abonado
                        ws=1
                        break
                    
            if abonado==0:
                contexto['abonado']=0
                #print("NO ENCONTRO EL ABONO EN EL BANCO")
            contexto["ws"]=ws                    
        if archivo7:
            print("ES VALIDO ARCHIVO 7")
            
            abono_dolar=pd.read_excel(archivo7,usecols=[1,2,3,4,5,6,7,8,9,10], nrows=1500)   # transbank         
            erp=pd.read_excel(archivo1,usecols=[1,2,3,4,5,6,7,8,9,10], nrows=1500)  # erp
             
            for fila in abono_dolar.index:
                c0='' # codigo autorizacion
                c1='' # documento
                c2='' # monto abono
                c3='' # monto transbank
                c4='' # diferencia
                c5='' # fecha
                c6='' # tipo tarjeta
                c7='' # status
                c8='' # observacion 
                c9='' # extra
                               
                i+= 1                
                texto = f"{i} - {abono_dolar.at[fila, 'Unnamed: 3']} - {abono_dolar.at[fila, 'Unnamed: 6']} - {abono_dolar.at[fila,'Unnamed: 7']}"
                #print(texto)
                tipo_tarjeta = abono_dolar.at[fila, 'Unnamed: 3']                
                monto_original = abono_dolar.at[fila, 'Unnamed: 6']
                codigo_autorizacion = abono_dolar.at[fila, 'Unnamed: 7']
                fecha_venta = abono_dolar.at[fila, 'Unnamed: 2']
                
                if pd.isna(codigo_autorizacion) or not tipo_tarjeta in ["AX","DI","VI","MC"]: 
                    continue
                
                erp['Unnamed: 9'] = erp['Unnamed: 9'].fillna(0)
                erp['Unnamed: 9'] = erp['Unnamed: 9'].astype(str).str.strip()
                find_erp = erp[erp['Unnamed: 9'].str.contains(str(codigo_autorizacion), na=False,case=False)]
                
                cantidad_filas = len(find_erp)
                #print(f"<<<<<<< Cantidad de filas encontradas para el codigo {codigo_autorizacion} en ERP: {cantidad_filas}")
                c0=codigo_autorizacion
                
                if cantidad_filas>0:
                    #print(f"          Encontrado en ERP el codigo: {codigo_autorizacion} con valor de {monto_original}")
                    c2=monto_original
                    
                    codigo_erp = find_erp.iloc[0]['Unnamed: 3']
                    valor_erp = abs(int(find_erp.iloc[0]['Unnamed: 8']))                                                        
                    
                    if "DI" in tipo_tarjeta:
                        #print("Buscando en Dinners US$----------->")                    
                        dinners = pd.read_excel(archivo3,usecols=[1,2,3,4,5,6,7,8,9,10,11,12,13,14], nrows=2500,engine='xlrd')
                        find_dinners = dinners[dinners['Unnamed: 2'] == codigo_erp]
                        c1=codigo_erp
                        
                        if len(find_dinners)>0:
                            #print(f"        Cantidad de filas encontradas para el codigo {codigo_erp} en Dinners: {len(find_dinners)} **************")                            
                            #print(f"        Encontrado en ERP y Dinners el codigo: {codigo_erp} con valor de: {monto_original} *************")
                            index=0
                            encontro = 0
                            for fila_dinners in find_dinners.index:
                                index += 1
                                texto_dinners = f"        {index} -Documento: {find_dinners.at[fila_dinners, 'Unnamed: 2']} - Otra Moneda: {find_dinners.at[fila_dinners, 'Unnamed: 12']} - Saldo: {find_dinners.at[fila_dinners,'Unnamed: 13']} - Saldo Corregido:{find_dinners.at[fila_dinners,'Unnamed: 14']}"
                                #print(texto_dinners)
                                otra_moneda = find_dinners.at[fila_dinners, 'Unnamed: 12']
                                saldo = find_dinners.at[fila_dinners, 'Unnamed: 13']
                                saldo_corregido = find_dinners.at[fila_dinners, 'Unnamed: 14']
                                if monto_original== otra_moneda:
                                    #print(f"     El monto original {monto_original} coincide con OTRA MONEDA {otra_moneda}")
                                    c7=f"El monto original {monto_original} coincide con OTRA MONEDA {otra_moneda}"
                                    c3=otra_moneda
                                        
                                if saldo==valor_erp:
                                    #print(f"     El valor ERP {valor_erp} coincide con SALDO {saldo}") 
                                    c8=f"El valor ERP {valor_erp} coincide con SALDO {saldo}"
                                    
                                    mo=monto_original
                                    om=otra_moneda                                   
                                    encontro = 1
                                    break
                                elif valor_erp==saldo_corregido:
                                    #print(f"     El valor ERP {valor_erp} coincide con SALDO CORREGIDO {saldo_corregido}")
                                    c8=f"El valor ERP {valor_erp} coincide con SALDO CORREGIDO {saldo_corregido}"
                                    mo=monto_original
                                    om=otra_moneda
                                    encontro = 1
                                    break
                            if encontro==0:
                                #print(f"\033[31m      NO ENCONTRO EL MONTO ORIGINAL {monto_original} EN NINGUNA COLUMNA DE DINNERS \033[0m")
                                c7=f"NO ENCONTRO EL MONTO ORIGINAL {monto_original} EN NINGUNA COLUMNA DE DINNERS"            
                            else:
                                diferencia = mo - om 
                                #print(f"     Diferencia entre monto original y otra moneda: {diferencia:.2f}")
                                c4=diferencia 
                                c3=om  
                        reg={"codigo_autorizacion":c0,"documento":c1,"monto_original":c2,"monto_transbank":c3,"diferencia":c4,"fecha_venta":fecha_venta,"tipo_tarjeta":tipo_tarjeta,"status":c7,"observacion":c8,"extra":"OK"}
                        informacion_proceso.append(reg)
                                                
                    if "VI" in tipo_tarjeta:
                        #print("Buscando en Visa US$----------------------->")                    
                        dinners = pd.read_excel(archivo4,usecols=[1,2,3,4,5,6,7,8,9,10,11,12,13,14], nrows=2500,engine='xlrd')
                        find_dinners = dinners[dinners['Unnamed: 2'] == codigo_erp]
                        c1=codigo_erp
                        
                        if len(find_dinners)>0:
                            #print(f"     Cantidad de filas encontradas para el codigo {codigo_erp} en Visa: {len(find_dinners)} **************")                            
                            #print(f"     Encontrado en ERP y Visa el codigo: {codigo_erp} con valor de {monto_original} *************")
                            index=0
                            encontro = 0
                            for fila_dinners in find_dinners.index:
                                index += 1
                                texto_dinners = f"     {index} -Documento: {find_dinners.at[fila_dinners, 'Unnamed: 2']} - Otra Moneda: {find_dinners.at[fila_dinners, 'Unnamed: 12']} - Saldo: {find_dinners.at[fila_dinners,'Unnamed: 13']} - Saldo Corregido:{find_dinners.at[fila_dinners,'Unnamed: 14']}"
                                #print(texto_dinners)
                                otra_moneda = find_dinners.at[fila_dinners, 'Unnamed: 12']
                                saldo = find_dinners.at[fila_dinners, 'Unnamed: 13']
                                saldo_corregido = find_dinners.at[fila_dinners, 'Unnamed: 14']
                                if monto_original== otra_moneda:
                                    #print(f"     El monto original {monto_original} coincide con OTRA MONEDA {otra_moneda}")
                                    c7=f"El monto original {monto_original} coincide con OTRA MONEDA {otra_moneda}"
                                    c3=otra_moneda
                                        
                                if saldo==valor_erp:
                                    #print(f"     El valor ERP {valor_erp} coincide con SALDO {saldo}") 
                                    c8=f"El valor ERP {valor_erp} coincide con SALDO {saldo}"
                                    
                                    mo=monto_original
                                    om=otra_moneda                                   
                                    encontro = 1
                                    break
                                elif valor_erp==saldo_corregido:
                                    #print(f"     El valor ERP {valor_erp} coincide con SALDO CORREGIDO {saldo_corregido}")
                                    c8=f"El valor ERP {valor_erp} coincide con SALDO CORREGIDO {saldo_corregido}"
                                    mo=monto_original
                                    om=otra_moneda
                                    encontro = 1
                                    break
                            if encontro==0:
                                #print(f"\033[31m      NO ENCONTRO EL MONTO ORIGINAL {monto_original} EN NINGUNA COLUMNA DE VISA \033[0m")            
                                c7=f"NO ENCONTRO EL MONTO ORIGINAL {monto_original} EN NINGUNA COLUMNA DE VISA"
                            else:
                                diferencia = mo - om 
                                #print(f"     Diferencia entre monto original y otra moneda: {diferencia:.2f}")   
                                c4=diferencia 
                                c3=om 
                        reg={"codigo_autorizacion":c0,"documento":c1,"monto_original":c2,"monto_transbank":c3,"diferencia":c4,"fecha_venta":fecha_venta,"tipo_tarjeta":tipo_tarjeta,"status":c7,"observacion":c8,"extra":"OK"}
                        informacion_proceso.append(reg)                                                        

                    if "MC" in tipo_tarjeta:
                        #print("Buscando en Master Card US$----------------------->")                    
                        dinners = pd.read_excel(archivo5,usecols=[1,2,3,4,5,6,7,8,9,10,11,12,13,14], nrows=2500,engine='xlrd')
                        find_dinners = dinners[dinners['Unnamed: 2'] == codigo_erp]
                        c1=codigo_erp
                        
                        if len(find_dinners)>0:
                            #print(f"     Cantidad de filas encontradas para el codigo {codigo_erp} en Visa: {len(find_dinners)} **************")                            
                            #print(f"     Encontrado en ERP y Master Card el codigo: {codigo_erp} con valor de {monto_original} *************")
                            index=0
                            encontro = 0
                            for fila_dinners in find_dinners.index:
                                index += 1
                                texto_dinners = f"     {index} -Documento: {find_dinners.at[fila_dinners, 'Unnamed: 2']} - Otra Moneda: {find_dinners.at[fila_dinners, 'Unnamed: 12']} - Saldo: {find_dinners.at[fila_dinners,'Unnamed: 13']} - Saldo Corregido:{find_dinners.at[fila_dinners,'Unnamed: 14']}"
                                #print(texto_dinners)
                                otra_moneda = find_dinners.at[fila_dinners, 'Unnamed: 12']
                                saldo = find_dinners.at[fila_dinners, 'Unnamed: 13']
                                saldo_corregido = find_dinners.at[fila_dinners, 'Unnamed: 14']
                                if monto_original== otra_moneda:
                                    #print(f"     El monto original {monto_original} coincide con OTRA MONEDA {otra_moneda}")
                                    c7=f"El monto original {monto_original} coincide con OTRA MONEDA {otra_moneda}"
                                    c3=otra_moneda
                                        
                                if saldo==valor_erp:
                                    #print(f"     El valor ERP {valor_erp} coincide con SALDO {saldo}") 
                                    c8=f"El valor ERP {valor_erp} coincide con SALDO {saldo}"
                                    
                                    mo=monto_original
                                    om=otra_moneda                                   
                                    encontro = 1
                                    break
                                elif valor_erp==saldo_corregido:
                                    #print(f"     El valor ERP {valor_erp} coincide con SALDO CORREGIDO {saldo_corregido}")
                                    c8=f"El valor ERP {valor_erp} coincide con SALDO CORREGIDO {saldo_corregido}"
                                    mo=monto_original
                                    om=otra_moneda
                                    encontro = 1
                                    break
                            if encontro==0:
                                #print(f"\033[31m      NO ENCONTRO EL MONTO ORIGINAL {monto_original} EN NINGUNA COLUMNA DE MASTER CARD \033[0m")            
                                c7=f"NO ENCONTRO EL MONTO ORIGINAL {monto_original} EN NINGUNA COLUMNA DE MASTER CARD"
                            else:
                                diferencia = mo - om 
                                #print(f"     Diferencia entre monto original y otra moneda: {diferencia:.2f}")   
                                c4=diferencia 
                                c3=om 
                        reg={"codigo_autorizacion":c0,"documento":c1,"monto_original":c2,"monto_transbank":c3,"diferencia":c4,"fecha_venta":fecha_venta,"tipo_tarjeta":tipo_tarjeta,"status":c7,"observacion":c8,"extra":"OK"}
                        informacion_proceso.append(reg)                                                        

                    if "AX" in tipo_tarjeta:
                        print("Buscando en American Express US$----------------------->")                    
                        dinners = pd.read_excel(archivo2,usecols=[1,2,3,4,5,6,7,8,9,10,11,12,13,14], nrows=2500,engine='xlrd')
                        find_dinners = dinners[dinners['Unnamed: 2'] == codigo_erp]
                        c1=codigo_erp
                        
                        if len(find_dinners)>0:
                            print(f"     Cantidad de filas encontradas para el codigo {codigo_erp} en Visa: {len(find_dinners)} **************")                            
                            print(f"     Encontrado en ERP y American Express el codigo: {codigo_erp} con valor de {monto_original} *************")
                            index=0
                            encontro = 0
                            for fila_dinners in find_dinners.index:
                                index += 1
                                texto_dinners = f"     {index} -Documento: {find_dinners.at[fila_dinners, 'Unnamed: 2']} - Otra Moneda: {find_dinners.at[fila_dinners, 'Unnamed: 12']} - Saldo: {find_dinners.at[fila_dinners,'Unnamed: 13']} - Saldo Corregido:{find_dinners.at[fila_dinners,'Unnamed: 14']}"
                                print(texto_dinners)
                                otra_moneda = find_dinners.at[fila_dinners, 'Unnamed: 12']
                                saldo = find_dinners.at[fila_dinners, 'Unnamed: 13']
                                saldo_corregido = find_dinners.at[fila_dinners, 'Unnamed: 14']
                                if monto_original== otra_moneda:
                                    print(f"     El monto original {monto_original} coincide con OTRA MONEDA {otra_moneda}")
                                    c7=f"El monto original {monto_original} coincide con OTRA MONEDA {otra_moneda}"
                                    c3=otra_moneda
                                        
                                if saldo==valor_erp:
                                    print(f"     El valor ERP {valor_erp} coincide con SALDO {saldo}") 
                                    c8=f"El valor ERP {valor_erp} coincide con SALDO {saldo}"
                                    
                                    mo=monto_original
                                    om=otra_moneda                                   
                                    encontro = 1
                                    break
                                elif valor_erp==saldo_corregido:
                                    print(f"     El valor ERP {valor_erp} coincide con SALDO CORREGIDO {saldo_corregido}")
                                    c8=f"El valor ERP {valor_erp} coincide con SALDO CORREGIDO {saldo_corregido}"
                                    mo=monto_original
                                    om=otra_moneda
                                    encontro = 1
                                    break
                            if encontro==0:
                                print(f"\033[31m      NO ENCONTRO EL MONTO ORIGINAL {monto_original} EN NINGUNA COLUMNA DE AMERICAN EXPRESS \033[0m")            
                                c7=f"NO ENCONTRO EL MONTO ORIGINAL {monto_original} EN NINGUNA COLUMNA DE AMERICAN EXPRESS"
                            else:
                                diferencia = mo - om 
                                print(f"     Diferencia entre monto original y otra moneda: {diferencia:.2f}")   
                                c4=diferencia 
                                c3=om 
                        reg={"codigo_autorizacion":c0,"documento":c1,"monto_original":c2,"monto_transbank":c3,"diferencia":c4,"fecha_venta":fecha_venta,"tipo_tarjeta":tipo_tarjeta,"status":c7,"observacion":c8,"extra":"OK"}
                        informacion_proceso.append(reg)                                                        
                else:
                    #print(f"\033[31m NO ENCONTRO EL CODIGO : {codigo_autorizacion} CUYO VALOR ES DE {monto_original}  \033[0m")
                    c7=f"NO ENCONTRO EL CODIGO : {codigo_autorizacion} CUYO VALOR ES DE {monto_original} "
                    c6=tipo_tarjeta
                    c0=codigo_autorizacion
                    c2=monto_original
                    c5=fecha_venta
                    c9="NO"
                    reg={"codigo_autorizacion":c0,"documento":c1,"monto_original":c2,"monto_transbank":c3,"diferencia":c4,"fecha_venta":fecha_venta,"tipo_tarjeta":tipo_tarjeta,"status":c7,"observacion":c8,"extra":c9}   
                    informacion_proceso.append(reg)
                    
        print("fin proceso ERP con Transbank")
        print("-----------------------------------") 
           
        if archivo1:
            # Lectura de ERP
            print("ES VALIDO")
            x=pd.read_excel(archivo1,usecols=[1,2,3,4,5,6,7,8,9,10], nrows=1500)
            # Convertir el DataFrame a una lista de listas
            #lista_de_listas = x.values.tolist()
            # Imprimir la lista de listas
            #print(lista_de_listas)
            
            print("Valores Archivo ERP -----------------------------------")            
            # Definir el valor que deseas buscar
            # valor_buscado = 1704139

            # Filtrar el DataFrame para encontrar las filas donde la columna 3 coincide con el valor
            # Nota: La "columna 3" cargada tiene el índice de Python 2 (la tercera posición).
            # Usaremos .iloc para referirnos a la posición numérica de la columna.

            # filas_encontradas_df = x[x.iloc[:, 2] == valor_buscado]
            # print(f"Fila encontradas:{filas_encontradas_df}")
            # print(f"Se encontraron {len(filas_encontradas_df)} filas con el valor '{valor_buscado}' en la columna 3.")
            # print(filas_encontradas_df)
            # posicion = 0
            # for fila in lista_de_listas:
            #     print(f"Posicion: {posicion} Fila: {fila}")
            #     posicion += 1
            
            # return HttpResponse("Archivo procesado correctamente.")
            diccionario = {}
            i=0
            for fila in x.index:
                
                i+= 1
                texto = f"{i} - {x.at[fila, 'Unnamed: 3']} - {x.at[fila, 'Unnamed: 5']} - {x.at[fila,'Unnamed: 8']}"
                # busca los textos que contengan US$ (DOLARES)

                if "US$" in texto:                    
                    #print(texto)
                    codigo_buscar = x.at[fila, 'Unnamed: 3']
                    valor_buscado = abs(int(x.at[fila, 'Unnamed: 8']))
                    
                    fecha_erp = x.at[fila, 'Unnamed: 10']
                    if (x.at[fila, 'Unnamed: 8'])<0:
                        # totales por tipo de tarjeta (TC = Tipo de Cuenta)
                        if "Amex US$" in texto:
                            dic_amex.append({"codigo":codigo_buscar,"valor":valor_buscado,"ERP":x.at[fila, 'Unnamed: 8']})
                            suma_amex = suma_amex + abs(int(x.at[fila, 'Unnamed: 8']))
                            
                        if "Dinners US$" in texto:
                            suma_dinners = suma_dinners + abs(int(x.at[fila, 'Unnamed: 8']))
                            
                        if "Master Card US$" in texto:
                            cant_master_card = cant_master_card +1
                            suma_master_card = suma_master_card + abs(int(x.at[fila, 'Unnamed: 8']))
                            
                        if "Visa US$" in texto:
                            suma_visa = suma_visa + abs(int(x.at[fila, 'Unnamed: 8']))
                        
                    #print(f"CODIGO A BUSCAR: {codigo_buscar} - Valor Buscado:{valor_buscado} ")
                    #print("-----------------------------------")
                    if "Amex US$" in texto:
                        cantidad=0
                        suma = 0
                        if archivo2:
                            #print("Archivo de AMEX US$")
                            amex=pd.read_excel(archivo2,usecols=[1,2,3,4,5,6,7,8,9,10,11,12,13,14], nrows=2500)
                            i_amex=0
                            #print("Valores Archivo AMEX US$ -----------------------------------")
                            sw=0
                            for fila_amex in amex.index:
                                codigo = amex.at[fila_amex, 'Unnamed: 2']
                                i_amex = i_amex +1
                                texto_amex = f"{i_amex} - {amex.at[fila_amex, 'Unnamed: 2']} - {amex.at[fila_amex, 'Unnamed: 12']} - {amex.at[fila_amex,'Unnamed: 13']} - {amex.at[fila_amex,'Unnamed: 14']}"
                                if codigo_buscar==codigo:
                                    cantidad = cantidad +1
                                    #print("----> Encontrada la Cuenta:")
                                    sw_e = 0
                                    if not math.isnan(amex.at[fila_amex, 'Unnamed: 12']): 
                                        if valor_buscado==abs(int(amex.at[fila_amex, 'Unnamed: 12'])):
                                           # print(f"Valor Buscado: {valor_buscado} esta presente en OTRA MONEDA {amex.at[fila_amex, 'Unnamed: 12']}")
                                            sw_e = 1
                                            suma = suma + abs(int(amex.at[fila_amex, 'Unnamed: 12']))
                                            
                                    if not math.isnan(amex.at[fila_amex, 'Unnamed: 13']):                                     
                                        if valor_buscado==abs(int(amex.at[fila_amex, 'Unnamed: 13'])):
                                            #print(f"Valor Buscado: {valor_buscado} esta presente en SALDO {amex.at[fila_amex, 'Unnamed: 13']}")
                                            sw_e = 1
                                            suma = suma + abs(int(amex.at[fila_amex, 'Unnamed: 13']))
                                            
                                    if not math.isnan(amex.at[fila_amex, 'Unnamed: 14']): 
                                        if valor_buscado==abs(int(amex.at[fila_amex, 'Unnamed: 14'])):
                                            #print(f"Valor Buscado: {valor_buscado} esta presente en SALDO CORREGIDO {amex.at[fila_amex, 'Unnamed: 14']}")
                                            sw_e = 1
                                            suma = suma + abs(int(amex.at[fila_amex, 'Unnamed: 14'])) 
                                    if sw_e == 0:
                                        print(f"\033[31m NO ESTA PRESENTE EL VALOR EN NINGUNA COLUMNA (amex)...CODIGO {codigo_buscar}..Valor:{valor_buscado}.. \033[0m")                                                                               
                                    #print(texto_amex)
                                    sw=1
                                    diccionario[f"{i}"]={"tarjeta":"Amex US$","codigo":abs(int(codigo_buscar)),"Valor":valor_buscado,"Status":"encontrado"}
                                    continue
                            if(sw==0):
                                #print(f"NO ENCONTRO EL CODIGO : {codigo_buscar} CUYO VALOR ES DE {valor_buscado}")
                                tc_inf.append("Amex US$")
                                documento_inf.append(abs(int(codigo_buscar)))
                                monto_buscar_inf.append(valor_buscado)
                                fecha_inf.append(fecha_erp)
                                diccionario[f"{i}"]={"tarjeta":"Amex US$","codigo":abs(int(codigo_buscar)),"Valor":valor_buscado,"Status":"no encontrado"}                                               
                    
                    # if "Dinners US$" in texto:
                    #     cantidad=0
                    #     suma = 0
                    #     if archivo3:
                    #         print("Archivo de Dinners US$")
                    #         dinners=pd.read_excel(archivo3,usecols=[1,2,3,4,5,6,7,8,9,10,11,12,13,14], nrows=2500)
                    #         i_dinners=0
                    #         print("Valores Archivo DINNERS US$ -----------------------------------")
                    #         sw=0
                    #         for fila_dinners in dinners.index:
                    #             codigo = dinners.at[fila_dinners, 'Unnamed: 2']
                    #             i_dinners = i_dinners +1
                    #             texto_dinners = f"{i_dinners} - {dinners.at[fila_dinners, 'Unnamed: 2']} - {dinners.at[fila_dinners, 'Unnamed: 12']} - {dinners.at[fila_dinners,'Unnamed: 13']} - {dinners.at[fila_dinners,'Unnamed: 14']}"
                    #             if codigo_buscar==codigo:
                    #                 cantidad = cantidad +1
                    #                 print("----> Encontrada la Cuenta:")
                    #                 sw_e = 0
                    #                 if not math.isnan(dinners.at[fila_dinners, 'Unnamed: 12']): 
                    #                     if valor_buscado==abs(int(dinners.at[fila_dinners, 'Unnamed: 12'])):
                    #                         print(f"Valor Buscado: {valor_buscado} esta presente en OTRA MONEDA {dinners.at[fila_dinners, 'Unnamed: 12']}")
                    #                         sw_e = 1
                    #                         suma = suma + abs(int(dinners.at[fila_dinners, 'Unnamed: 12']))                                            
                                            
                    #                 if not math.isnan(dinners.at[fila_dinners, 'Unnamed: 13']):                                     
                    #                     if valor_buscado==abs(int(dinners.at[fila_dinners, 'Unnamed: 13'])):
                    #                         print(f"Valor Buscado: {valor_buscado} esta presente en SALDO {dinners.at[fila_dinners, 'Unnamed: 13']}")
                    #                         sw_e = 1
                    #                         suma = suma + abs(int(dinners.at[fila_dinners, 'Unnamed: 13']))
                                            
                    #                 if not math.isnan(dinners.at[fila_dinners, 'Unnamed: 14']): 
                    #                     if valor_buscado==abs(int(dinners.at[fila_dinners, 'Unnamed: 14'])):
                    #                         print(f"Valor Buscado: {valor_buscado} esta presente en SALDO CORREGIDO {dinners.at[fila_dinners, 'Unnamed: 14']}")
                    #                         sw_e = 1
                    #                         suma = suma + abs(int(dinners.at[fila_dinners, 'Unnamed: 14']))
                    #                 if sw_e == 0:
                    #                     print(f"NO ESTA PRESENTE EL VALOR EN NINGUNA COLUMNA (dinners)...CODIGO {codigo_buscar}..Valor:{valor_buscado}..")                                    
                    #                 print(texto_dinners)
                    #                 sw=1
                    #                 diccionario[f"{i}"]={"tarjeta":"Dinners US$","codigo":abs(int(codigo_buscar)),"Valor":valor_buscado,"Status":"encontrado"}
                    #         if(sw==0):
                    #             print(f"NO ENCONTRO EL CODIGO : {codigo_buscar} CUYO VALOR ES DE {valor_buscado}")
                    #             tc_inf.append("Dinners US$")
                    #             documento_inf.append(abs(int(codigo_buscar)))
                    #             monto_buscar_inf.append(valor_buscado)
                    #             fecha_inf.append(fecha_erp)
                    #             diccionario[f"{i}"]={"tarjeta":"Dinners US$","codigo":abs(int(codigo_buscar)),"Valor":valor_buscado,"Status":"no encontrado"}                                             
                    
                    
                    if "Master Card US$" in texto:
                        # por cada codigo que busco en el ERP voy a buscar de forma ciclica en el archivo de mastercard 1 a 1 
                        #print(f"Buscando en Master Card el codigo: {codigo_buscar} con valor de {valor_buscado}")
                        cantidad=0
                        if archivo5:
                            #print("Archivo de Master Card US$")
                            mc=pd.read_excel(archivo5,usecols=[1,2,3,4,5,6,7,8,9,10,11,12,13,14], nrows=2500)
                            i_mc=0
                            #print("Valores Archivo Master Card US$ -----------------------------------")
                            sw=0
                            sw_e = 0
                            om = 0
                            sal = 0
                            sal_corr = 0    
                                                            
                            filtro = mc['Unnamed: 2'] == codigo_buscar
                            
                            if not filtro.any():
                                # no encontro el codigo en mastercard
                                #print(f"\033[31m NO ENCONTRO EL CODIGO : {codigo_buscar} CUYO VALOR ES DE {valor_buscado}  \033[0m")
                                tc_inf.append("Master Card US$")
                                documento_inf.append(abs(int(codigo_buscar)))
                                monto_buscar_inf.append(valor_buscado)
                                fecha_inf.append(fecha_erp)
                                diccionario[f"{i}"]={"tarjeta":"Master Card US$","codigo":abs(int(codigo_buscar)),"Valor":valor_buscado,"Status":"no encontrado"}                                             
                                continue
                            
                            #for fila_mc in mc.index:
                            cantidad_filas = len(mc.index[filtro])
                            
                            #print(f"Cantidad de filas encontradas para el codigo {codigo_buscar} en Master Card: {cantidad_filas}")
                            paso = 0
                            
                            for fila_mc in mc.index[filtro]:
                                paso = paso + 1
                                codigo = mc.at[fila_mc, 'Unnamed: 2'] # Numeros de Documentos recuperados desde archivo "Master Card"
                                if pd.isna(codigo):
                                    continue
                                i_mc = i_mc +1
                                texto_mc = f"{i_mc} -Documento: {mc.at[fila_mc, 'Unnamed: 2']} - Otra Moneda: {mc.at[fila_mc, 'Unnamed: 12']} - Saldo: {mc.at[fila_mc,'Unnamed: 13']} - Saldo Corregido:{mc.at[fila_mc,'Unnamed: 14']}"
                                #print(texto_mc)
                                # el codigo que busco del ERP es igual al codigo que estoy leyendo del archivo de mastercard
                                if codigo_buscar==codigo:
                                    cantidad = cantidad +1
                                    # print(f"----> Encontrada la Cuenta:{codigo}")
                                    # recupero de la fila actual los valores de las columnas 12,13,14 para ver si alguno coincide con el valor buscado                                    
                                    if not math.isnan(mc.at[fila_mc, 'Unnamed: 12']): 
                                        if valor_buscado==abs(int(mc.at[fila_mc, 'Unnamed: 12'])):
                                            #print(f"Valor Buscado: {valor_buscado} esta presente en OTRA MONEDA {mc.at[fila_mc, 'Unnamed: 12']}")
                                            sw_e = 1
                                            om = abs(int(mc.at[fila_mc, 'Unnamed: 12']))
                                            suma_mc = suma_mc + abs(int(mc.at[fila_mc, 'Unnamed: 12']))                                            
                                            
                                    if not math.isnan(mc.at[fila_mc, 'Unnamed: 13']):                                     
                                        if valor_buscado==abs(int(mc.at[fila_mc, 'Unnamed: 13'])):
                                            #print(f"Valor Buscado: {valor_buscado} esta presente en SALDO {mc.at[fila_mc, 'Unnamed: 13']}")
                                            sw_e = 1
                                            sal = abs(int(mc.at[fila_mc, 'Unnamed: 13']))
                                            suma_mc = suma_mc + abs(int(mc.at[fila_mc, 'Unnamed: 13']))
                                            
                                    if not math.isnan(mc.at[fila_mc, 'Unnamed: 14']): 
                                        if valor_buscado==abs(int(mc.at[fila_mc, 'Unnamed: 14'])):
                                            #print(f"Valor Buscado: {valor_buscado} esta presente en SALDO CORREGIDO {mc.at[fila_mc, 'Unnamed: 14']}")
                                            sw_e = 1
                                            sal_corr = abs(int(mc.at[fila_mc, 'Unnamed: 14']))
                                            suma_mc = suma_mc + abs(int(mc.at[fila_mc, 'Unnamed: 14']))
                                    #if sw_e == 0:
                                        #print(f"No ENCONTRO VALOR:{valor_buscado}")
                                    #    print(f"\033[31m NO ESTA PRESENTE EL VALOR EN NINGUNA COLUMNA (master card)...CODIGO {codigo_buscar}..Valor:{valor_buscado}..Otra Moneda.{om}..Saldo.{sal}..Saldo Corr.{sal_corr} \033[0m")
                                    #print(texto_mc)
                                    sw=1
                                    diccionario[f"{i}"]={"tarjeta":"Master Card US$","codigo":abs(int(codigo_buscar)),"Valor":valor_buscado,"Status":"encontrado"}
                            if(paso==cantidad_filas) and (sw_e==0):
                                #print(f"\033[31m NO ENCONTRO EL CODIGO : {codigo_buscar} CUYO VALOR ES DE {valor_buscado}  \033[0m")
                                tc_inf.append("Master Card US$")
                                documento_inf.append(abs(int(codigo_buscar)))
                                monto_buscar_inf.append(valor_buscado)
                                fecha_inf.append(fecha_erp)
                                diccionario[f"{i}"]={"tarjeta":"Master Card US$","codigo":abs(int(codigo_buscar)),"Valor":valor_buscado,"Status":"no encontrado"}                                             

                    # if "Visa US$" in texto:
                    #     cantidad=0
                    #     suma = 0
                    #     if archivo4:
                    #         print("Archivo de Visa US$")
                    #         mc=pd.read_excel(archivo4,usecols=[1,2,3,4,5,6,7,8,9,10,11,12,13,14], nrows=2500)
                    #         i_mc=0
                    #         print("Valores Archivo Visa US$ -----------------------------------")
                    #         sw=0
                    #         for fila_mc in mc.index:
                    #             codigo = mc.at[fila_mc, 'Unnamed: 2']
                    #             i_mc = i_mc +1
                    #             texto_mc = f"{i_mc} - {mc.at[fila_mc, 'Unnamed: 2']} - {mc.at[fila_mc, 'Unnamed: 12']} - {mc.at[fila_mc,'Unnamed: 13']} - {mc.at[fila_mc,'Unnamed: 14']}"
                    #             if codigo_buscar==codigo:
                    #                 cantidad = cantidad +1
                    #                 print("----> Encontrada la Cuenta:")
                    #                 sw_e = 0
                    #                 if not math.isnan(mc.at[fila_mc, 'Unnamed: 12']): 
                    #                     if valor_buscado==abs(int(mc.at[fila_mc, 'Unnamed: 12'])):
                    #                         print(f"Valor Buscado: {valor_buscado} esta presente en OTRA MONEDA {mc.at[fila_mc, 'Unnamed: 12']}")
                    #                         sw_e = 1
                    #                         suma = suma + abs(int(mc.at[fila_mc, 'Unnamed: 12']))                                            
                                            
                    #                 if not math.isnan(mc.at[fila_mc, 'Unnamed: 13']):                                     
                    #                     if valor_buscado==abs(int(mc.at[fila_mc, 'Unnamed: 13'])):
                    #                         print(f"Valor Buscado: {valor_buscado} esta presente en SALDO {mc.at[fila_mc, 'Unnamed: 13']}")
                    #                         sw_e = 1
                    #                         suma = suma + abs(int(mc.at[fila_mc, 'Unnamed: 13']))
                                            
                    #                 if not math.isnan(mc.at[fila_mc, 'Unnamed: 14']): 
                    #                     if valor_buscado==abs(int(mc.at[fila_mc, 'Unnamed: 14'])):
                    #                         print(f"Valor Buscado: {valor_buscado} esta presente en SALDO CORREGIDO {mc.at[fila_mc, 'Unnamed: 14']}")
                    #                         sw_e = 1
                    #                         suma = suma + abs(int(mc.at[fila_mc, 'Unnamed: 14']))
                    #                 if sw_e == 0:                                    
                    #                     print(f"NO ESTA PRESENTE EL VALOR EN NINGUNA COLUMNA (visa)...CODIGO {codigo_buscar}..Valor:{valor_buscado}..")
                    #                 print(texto_mc)
                    #                 sw=1
                    #                 diccionario[f"{i}"]={"tarjeta":"Visa US$","codigo":abs(int(codigo_buscar)),"Valor":valor_buscado,"Status":"encontrado"}
                    #         if(sw==0):
                    #             print(f"NO ENCONTRO EL CODIGO : {codigo_buscar} CUYO VALOR ES DE {valor_buscado}")
                    #             tc_inf.append("Visa US$")
                    #             documento_inf.append(abs(int(codigo_buscar)))
                    #             monto_buscar_inf.append(valor_buscado)
                    #             fecha_inf.append(fecha_erp)
                    #             diccionario[f"{i}"]={"tarjeta":"Visa US$","codigo":abs(int(codigo_buscar)),"Valor":valor_buscado,"Status":"no encontrado"}                                             

                continue
                clave = x.at[fila, 'Unnamed: 3']
                valor = x.at[fila, 'Unnamed: 6']
                    # Verifica que la clave no sea nan
                if pd.isna(clave):
                    # print("Clave nula, se omite:", clave)
                    continue
                if x.at[fila, 'Unnamed: 3'] in diccionario:
                    # print("La clave ya existe:", x.at[fila, 'Unnamed: 3'])
                    diccionario[x.at[fila, 'Unnamed: 3']] = diccionario[x.at[fila, 'Unnamed: 3']] + [x.at[fila, 'Unnamed: 6']]        
                else:
                    # print("La clave no existe, se crea:", x.at[fila, 'Unnamed: 3']) 
                    # print("Valor:", x.at[fila, 'Unnamed: 6']) 
                    # print("Tipo de valor:", type(x.at[fila, 'Unnamed: 6']))
                    if str(valor).isnumeric():
                        diccionario[str(clave)] = [valor]      
                    # diccionario[str(clave)].append(valor)
                
                print(texto)
                i += 1
                # for columna in x.columns:
                #     valor = x.at[fila, columna]
                #     print("columna", columna)
                #     print(f"Fila {fila}, Columna {columna}: {valor}")
        
        print("TERMINO PROCESO ERP")
        print("-----------------------------------")
        print(f"Total Master Card US$ Encontrado en ERP: {suma_mc}")
        
        if archivo7:
                #print("procesar Transbank con ERP (archivo7 con archivo1)")
                #print("ES VALIDO ARCHIVO 7")
                TRANS   =pd.read_excel(archivo7,usecols=[1,2,3,4,5,6,7,8,9,10], nrows=1500) #Archivo Transbank
                diccionario_2 = {}
                i2=0
                for fila in TRANS.index:
                    i2+= 1
                    fecha = TRANS.at[fila, 'Unnamed: 2']
                    tipo_tarjeta = TRANS.at[fila, 'Unnamed: 3']
                    monto_original = TRANS.at[fila, 'Unnamed: 6']
                    codigo_aut = TRANS.at[fila, 'Unnamed: 7']
                    texto = f"{i2} - {fecha} - {tipo_tarjeta} - {monto_original} - {codigo_aut}"
                    #print(f"TRANSBANK: -----> {texto}")
                    ERP=pd.read_excel(archivo1,usecols=[1,2,3,4,5,6,7,8,9,10], nrows=1500) #Archivo ERP
                    sw=0
                    for fila_erp in ERP.index:
                        monto =  ERP.at[fila_erp, 'Unnamed: 8']
                        codigo_erp = ERP.at[fila_erp, 'Unnamed: 9']
                        if not pd.isna(tipo_tarjeta):                            
                            if str(codigo_aut) in str(codigo_erp):
                                sw=1
                                #print(f"Encontrado ..... (monto original: {monto_original}) - (monto ERP: {monto})")                                
                                diccionario_2[f"{i2}"]={"tarjeta":tipo_tarjeta,"codigo":str(codigo_aut),"Valor":abs(int(monto)),"Status":"encontrado"}    
                    if sw==0:
                        #print("No se encontro ------")
                        diccionario_2[f"{i2}"]={"tarjeta":tipo_tarjeta,"codigo":(codigo_aut),"Valor":monto_original,"Status":"no encontrado"}
                
        diccionario_temp=[
                {'Amex US$': 'Amex US$', 'Valor US': 0,'Valor TBK':0,'Diferencia':0},
                {'Dinners US$':'Dinners US$', 'Valor US': 0,'Valor TBK':0,'Diferencia':0},
                {'Master Card US$': 'Master Card US$', 'Valor US': 0,'Valor TBK':0,'Diferencia':0},
                {'Visa US$': 'Visa US$', 'Valor US': 0,'Valor TBK':0,'Diferencia':0},
                {'Total': 'Total', 'Valor US': 0,'Valor TBK':0,'Diferencia':0}
            ]
        total_amex=0
        total_dinners=0
        total_mc=0
        total_visa=0
        for clave, valor in diccionario.items():
            #print(f"Clave: {clave} -> Valor: {valor} -> Dato Valor 1:{list(valor.values())[0]} - Valor: ")
            valores = list(valor.values()) 
            if valores[3]=="encontrado":               
                if "Amex US$" in valores[0]:
                    total_amex += valores[2] 
                if "Dinners US$" in valores[0]:
                    total_dinners += valores[2]
                if "Master Card US$" in valores[0]:
                    total_mc += valores[2]
                if "Visa US$" in valores[0]:
                    total_visa += valores[2]
        diccionario_temp[0]["Valor US"]=total_amex
        diccionario_temp[1]["Valor US"]=total_dinners
        diccionario_temp[2]["Valor US"]=total_mc
        diccionario_temp[3]["Valor US"]=total_visa
        diccionario_temp[4]["Valor US"]=total_visa+total_amex+total_dinners+total_mc
        #print(diccionario_temp)
        
        for clave, valor in diccionario_2.items():
            #print(f"Clave: {clave} -> Valor: {valor} -> Dato Valor 1:{list(valor.values())[0]} - Valor: ")
            valores = list(valor.values())
            #print(f"Clave:{clave} Valores:{valor}") 
            if valores[3]=="encontrado":               
                if "AX" in valores[0]:
                    total_amex += valores[2] 
                if "DI" in valores[0]:
                    total_dinners += valores[2]
                if "MC" in valores[0]:
                    total_mc += valores[2]
                if "VI" in valores[0]:
                    total_visa += valores[2]
                    
        diccionario_temp[0]["Valor TBK"]=total_amex
        diccionario_temp[1]["Valor TBK"]=total_dinners
        diccionario_temp[2]["Valor TBK"]=total_mc
        diccionario_temp[3]["Valor TBK"]=total_visa
        diccionario_temp[4]["Valor TBK"]=total_visa+total_amex+total_dinners+total_mc
        #print(diccionario_temp)
        diccionario_temp[0]["Diferencia"] = diccionario_temp[0]["Valor TBK"] - diccionario_temp[0]["Valor US"]
        diccionario_temp[1]["Diferencia"] = diccionario_temp[1]["Valor TBK"] - diccionario_temp[1]["Valor US"]
        diccionario_temp[2]["Diferencia"] = diccionario_temp[2]["Valor TBK"] - diccionario_temp[2]["Valor US"]
        diccionario_temp[3]["Diferencia"] = diccionario_temp[3]["Valor TBK"] - diccionario_temp[3]["Valor US"]
        diccionario_temp[4]["Diferencia"] = diccionario_temp[4]["Valor TBK"] - diccionario_temp[4]["Valor US"]
        
        contexto["data"]=diccionario_temp        
    else:
        print("es invalido")

    diccionario_dolar=[
        {'Amex US$': 'Amex US$', 'Valor US': 0,'Valor TBK':0,'Diferencia':0},
        {'Dinners US$':'Dinners US$', 'Valor US': 0,'Valor TBK':0,'Diferencia':0},
        {'Master Card US$': 'Master Card US$', 'Valor US': 0,'Valor TBK':0,'Diferencia':0},
        {'Visa US$': 'Visa US$', 'Valor US': 0,'Valor TBK':0,'Diferencia':0},
        {'Total': 'Total', 'Valor US': 0,'Valor TBK':0,'Diferencia':0}
    ]
    v_amex=round(sum(item['monto_original'] for item in informacion_proceso if item['tipo_tarjeta'] == 'AX' and item['extra']=='OK'),2)
    v_dinners=round(sum(item['monto_original'] for item in informacion_proceso if item['tipo_tarjeta'] == 'DI' and item['extra']=='OK'),2)
    v_mc=round(sum(item['monto_original'] for item in informacion_proceso if item['tipo_tarjeta'] == 'MC' and item['extra']=='OK'),2)
    v_visa=round(sum(item['monto_original'] for item in informacion_proceso if item['tipo_tarjeta'] == 'VI' and item['extra']=='OK'),2)
    suma_total=v_amex+v_dinners+v_mc+v_visa
    
    diccionario_dolar[0]["Valor US"]= round(sum(item['monto_original'] for item in informacion_proceso if item['tipo_tarjeta'] == 'AX' and item['extra']=='OK'),2)
    diccionario_dolar[1]["Valor US"]= round(sum(item['monto_original'] for item in informacion_proceso if item['tipo_tarjeta'] == 'DI' and item['extra']=='OK'),2)
    diccionario_dolar[2]["Valor US"]= round(sum(item['monto_original'] for item in informacion_proceso if item['tipo_tarjeta'] == 'MC' and item['extra']=='OK'),2)
    diccionario_dolar[3]["Valor US"]= round(sum(item['monto_original'] for item in informacion_proceso if item['tipo_tarjeta'] == 'VI' and item['extra']=='OK'),2)
    diccionario_dolar[4]["Valor US"]= round(suma_total,2)
    
    # Reemplazar la coma por punto y convertir a float
    for item in informacion_proceso:
        if isinstance(item['monto_transbank'], str):  # Solo si es una cadena
            item['monto_transbank'] = item['monto_transbank'].replace(',', '.')  # Reemplazar ',' por '.'
        try:
            item['monto_transbank'] = float(item['monto_transbank'])  # Convertir a float
        except ValueError:
            item['monto_transbank'] = 0.0  # Si no es convertible, asigna 0.0

    # informacion_proceso['monto_transbank'] = pd.to_numeric(informacion_proceso['monto_transbank'], errors='coerce')
    v_amex_tbk=round(sum(item['monto_transbank'] for item in informacion_proceso if item['tipo_tarjeta'] == 'AX' and item['extra']=='OK'),2)
    v_dinners_tbk=round(sum(item['monto_transbank'] for item in informacion_proceso if item['tipo_tarjeta'] == 'DI' and item['extra']=='OK'),2)
    v_mc_tbk=round(sum(item['monto_transbank'] for item in informacion_proceso if item['tipo_tarjeta'] == 'MC' and item['extra']=='OK'),2)
    v_visa_tbk=round(sum(item['monto_transbank'] for item in informacion_proceso if item['tipo_tarjeta'] == 'VI' and item['extra']=='OK'),2)
    suma_total_tbk=v_amex_tbk+v_dinners_tbk+v_mc_tbk+v_visa_tbk
    diccionario_dolar[4]["Valor TBK"]= round(suma_total_tbk,2)
    
    diccionario_dolar[0]["Valor TBK"]= round(sum(item['monto_transbank'] for item in informacion_proceso if item['tipo_tarjeta'] == 'AX' and item['extra']=='OK'),2)
    diccionario_dolar[1]["Valor TBK"]= round(sum(item['monto_transbank'] for item in informacion_proceso if item['tipo_tarjeta'] == 'DI' and item['extra']=='OK'),2)
    diccionario_dolar[2]["Valor TBK"]= round(sum(item['monto_transbank'] for item in informacion_proceso if item['tipo_tarjeta'] == 'MC' and item['extra']=='OK'),2)
    diccionario_dolar[3]["Valor TBK"]= round(sum(item['monto_transbank'] for item in informacion_proceso if item['tipo_tarjeta'] == 'VI' and item['extra']=='OK'),2)
    
    diccionario_dolar[0]["Diferencia"] = round(diccionario_dolar[0]["Valor TBK"] - diccionario_dolar[0]["Valor US"],2)
    diccionario_dolar[1]["Diferencia"] = round(diccionario_dolar[1]["Valor TBK"] - diccionario_dolar[1]["Valor US"],2)
    diccionario_dolar[2]["Diferencia"] = round(diccionario_dolar[2]["Valor TBK"] - diccionario_dolar[2]["Valor US"],2)
    diccionario_dolar[3]["Diferencia"] = round(diccionario_dolar[3]["Valor TBK"] - diccionario_dolar[3]["Valor US"],2)
    diccionario_dolar[4]["Diferencia"] = round(diccionario_dolar[4]["Valor TBK"] - diccionario_dolar[4]["Valor US"],2)
    
    total_diferencia=diccionario_dolar[0]["Diferencia"]+diccionario_dolar[1]["Diferencia"]+diccionario_dolar[2]["Diferencia"]+diccionario_dolar[3]["Diferencia"]
    diccionario_dolar[4]["Diferencia"]=round(total_diferencia,2)
       
    print("-----------------------------------")
    print("Total Amex US$ en ERP hasta ahora:",suma_amex)
    print("Total Dinners US$ en ERP hasta ahora:",suma_dinners)
    print("Total Master Card US$ en ERP hasta ahora:",suma_master_card, " Cantidad de Operaciones:",cant_master_card)
    print("Total Visa US$ en ERP hasta ahora:",suma_visa)
    print("Diccionario Amex:",dic_amex)    
    print("Fin Proceso de Archivos")
    print("-----------------------------------")
    
    contexto["data_dolar"]=diccionario_dolar
    #reg={"codigo_autorizacion":c0,"documento":c1,"monto_original":c2,"monto_transbank":c3,"diferencia":c4,
    # "fecha_venta":fecha_venta,"tipo_tarjeta":tipo_tarjeta,"status":c7,"observacion":c8}
    
    # print("Documentos no encontrados:")
    # for i in informacion_proceso:
    #     print("Codigo Autorizacion:",i['codigo_autorizacion']," Documento:",i["documento"]," Monto Original:",i['monto_original']," TC:",i['tipo_tarjeta']," Fecha Venta:",i['fecha_venta']," Status:",i['status']," Observacion:",i['observacion'])           
    return render(request,"procesamiento.html",contexto)


data = {}
def descargar_excel_ant(request):
    # Datos de ejemplo
    data = {
        'Documento': documento_inf,
        'Monto': monto_buscar_inf,
        'Fecha': fecha_inf,
        'TC':tc_inf
    }
    df = pd.DataFrame(data)


    # Crear archivo Excel en memoria
    buffer = BytesIO()
    df.to_excel(buffer, index=False, engine='openpyxl')
    buffer.seek(0)

    # Crear respuesta HTTP para descarga
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=mi_archivo.xlsx'
    return response


# descargar excel con titulo en la primera fila
def descargar_excel(request):
    # Datos de ejemplo
    data = {
        'Documento': documento_inf,
        'Monto': monto_buscar_inf,
        'Fecha': fecha_inf,
        'TC': tc_inf
    }
    df = pd.DataFrame(data)

    # Crear archivo Excel en memoria
    buffer = BytesIO()
    df.to_excel(buffer, index=False, engine='openpyxl', startrow=1)  # empezar en fila 2
    buffer.seek(0)

    # Abrir el workbook con openpyxl para agregar el título
    wb = load_workbook(buffer)
    ws = wb.active

    # Agregar título en la primera fila
    ws['A1'] = "Reporte de Transacciones No Encontradas"  # texto del título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=df.shape[1])  # fusionar celdas del título

    # Guardar nuevamente en el buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Crear respuesta HTTP para descarga
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=mi_archivo.xlsx'
    return response
   
   
# descargar excel con titulo en la primera fila
def descargar_excel_dif(request):

    df = pd.DataFrame(informacion_proceso)
    df = df[df['extra'] == 'OK']
    df.sort_values(by=['tipo_tarjeta','status'], inplace=True)
    mascara_con_valor = pd.notna(df['documento'])
    df[mascara_con_valor]
    # Crear archivo Excel en memoria
    buffer = BytesIO()
    df.to_excel(buffer, index=False, engine='openpyxl', startrow=1)  # empezar en fila 2
    buffer.seek(0)

    # Abrir el workbook con openpyxl para agregar el título
    wb = load_workbook(buffer)
    ws = wb.active

    # Agregar título en la primera fila
    ws['A1'] = "Reporte de Transacciones Abonadas y Diferencias"  # texto del título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=df.shape[1])  # fusionar celdas del título

    # Guardar nuevamente en el buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Crear respuesta HTTP para descarga
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=informacion_proceso.xlsx'
    return response
   
   # descargar excel con titulo en la primera fila
def descargar_excel_nopresente(request):

    df = pd.DataFrame(informacion_proceso)
    df = df[df['extra'] == 'NO']
    df.sort_values(by=['tipo_tarjeta','status'], inplace=True)
    mascara_con_valor = pd.notna(df['documento'])
    df[mascara_con_valor]
    df = df.drop_duplicates()
    # Crear archivo Excel en memoria
    buffer = BytesIO()
    df.to_excel(buffer, index=False, engine='openpyxl', startrow=1)  # empezar en fila 2
    buffer.seek(0)

    # Abrir el workbook con openpyxl para agregar el título
    wb = load_workbook(buffer)
    ws = wb.active

    # Agregar título en la primera fila
    ws['A1'] = "Reporte de Transacciones No Encontradas"  # texto del 
    ws['A2']="(Transacciones que no se encuentran en el ERP)"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=df.shape[1])  # fusionar celdas del título

    # Guardar nuevamente en el buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Crear respuesta HTTP para descarga
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=transacciones_no_ubicadas.xlsx'
    return response
   
   
   
   
def generar_qr2(req):
    # Datos a codificar
    data = "https://example.com"
    
    # Crear código QR
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(data)
    qr.make(fit=True)

    img = qr.make_image(fill='black', back_color='white')

    # Convertir a bytes
    buf = BytesIO()
    img.save(buf, format='PNG')
    img_bytes = buf.getvalue()

    # Guardar
    qr_image = ContentFile(buf.getvalue(), 'qr_code.png')
    # Responder con imagen
    return HttpResponse(img_bytes, content_type='image/png')
     
def login(request):
    contexto = {}
    if request.POST:
        nombre = request.POST.get("email")
        password = request.POST.get("pass")
        us = authenticate(request,username=nombre,password=password)
        if us is not None and us.is_active:
            login_aut(request,us)
            usuario=User.objects.get(email=request.POST.get("email"))
            print("usuario")
            print(usuario.id)
            cli=Cliente.objects.get(email=nombre)
            request.session["email"]=nombre
            return render(request,"index.html",contexto)
        else:
            contexto = {"mensaje":"usuario y contraseña incorrecto"}
            return render(request,"login.html",contexto)        
    return render(request,"login.html",contexto)

def enviar_codigo_qr(request,clave,correo):
    
    # Datos que quieres convertir en un código QR
    data = clave  # Reemplázalo con la URL o información que desees

    # Generar el código QR
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(data)
    qr.make(fit=True)

    # Crear la imagen del código QR
    img = qr.make_image(fill_color="black", back_color="white")

    # Guardar la imagen en un buffer
    buffer = BytesIO()
    img.save(buffer, format="PNG")
    buffer.seek(0)

    # Crear el mensaje de correo electrónico
    email = EmailMessage(
        subject=correo,
        body='Adjunto encontrarás tu código QR.de tu reserva',
        from_email='fm_campos@yahoo.com',  # Reemplaza con tu correo
        to=[correo],   # Reemplaza con el correo del destinatario
    )

    # Adjuntar la imagen del código QR
    email.attach('codigo_qr.png', buffer.getvalue(), 'image/png')

    # Enviar el correo
    email.send()

    comentarios=Comentario.objects.all()
    contexto={'comentarios':comentarios}
    contexto["mensaje"]="OK"
    return 1

def cerrar_sesion(request):
    contex = {}
    logout(request)
    return render(request,"index.html",contex)
    
